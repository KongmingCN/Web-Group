#! python3
import openpyxl
filename = "scores.xlsx"

wb = openpyxl.load_workbook(filename)
sheet1 = wb.get_sheet_by_name('Sheet1')
sheet2 = wb.get_sheet_by_name('Sheet2')


for i in range(2,sheet2.max_row+1):
    current_id = sheet2.cell(column = 2 ,row = i).value.replace(' ','')
    current_score = sheet2.cell(column = 3,row = i).value
    for j in range(2,sheet1.max_row+1):
        if sheet1.cell(column = 2,row = j).value.replace(' ','') == current_id:
            sheet1.cell(column = 3,row = j).value = current_score
            break

wb.save('scores.xlsx')

            
