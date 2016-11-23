import openpyxl

wb = openpyxl.load_workbook('scores.xlsx')
sheet1 = wb.get_sheet_by_name('Sheet1')
sheet2 = wb.get_sheet_by_name('Sheet2')

dic2 = {}
for i in range(sheet2.max_row-1):
    dic2[sheet2['B{}'.format(i+2)].value.replace('\t', '').replace(' ', '')] = sheet2['C{}'.format(i+2)].value

print(dic2)

for i in range(sheet1.max_row-1):
    try:
        sheet1['C{}'.format(i+2)].value = dic2[sheet1['B{}'.format(i+2)].value.replace('\t', '').replace(' ','')]
        print(str(i+2)+'OK')
    except AttributeError:
        print(str(i+2)+'error')

wb.save('scores.xlsx')

print('done')
